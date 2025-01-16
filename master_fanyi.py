from openpyxl.styles import Font, Color, PatternFill,Alignment,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook,load_workbook
import pandas as pd

wb = Workbook()
ws1 = wb.active
ws1.title = "Sample"
alignment = Alignment(vertical='center', wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
master_Sample = pd.read_excel("./master/master.xlsx", sheet_name="Sample")
df_Sample = pd.read_excel("./excel_file/HK_4096786-2.xlsx", sheet_name="Sample Details")
df_Sample=df_Sample.fillna("")
dict1=master_Sample.to_dict()
keys=list(dict1["Sample_eng"].values())
values=list(dict1["Sample_cn"].values())
dict2 = {key: value for key, value in zip(keys, values)}

check_point=["Sample Type","Sample Personalization","Shipping Method","City State Zip","Special Instructions"]
Sample_group=df_Sample.groupby(by=check_point)
if Sample_group.groups.__len__()>2:  #有多個不同的sample翻譯
    ws1.append(df_Sample.columns.tolist())
    for index, row in df_Sample.iterrows():
        Lot=row["Lot"]
        Sample_Type_eng=row["Sample Type"]
        Sample_Personalization_eng=row["Sample Personalization"]
        Shipping_Method_eng=row["Shipping Method"]
        City_State_Zip_eng=row["City State Zip"]
        Special_Instructions_eng=row["Special Instructions"]

        Sample_Type = Sample_Type_eng+'\n'+dict2.get(Sample_Type_eng,"")
        Sample_Personalization = Sample_Personalization_eng+'\n'+dict2.get(Sample_Personalization_eng,"")
        Shipping_Method = Shipping_Method_eng+'\n'+dict2.get(Shipping_Method_eng,"")
        City_State_Zip = City_State_Zip_eng+'\n'+dict2.get(City_State_Zip_eng,"")
        Special_Instructions = Special_Instructions_eng+'\n'+dict2.get(Special_Instructions_eng,"")
        ws1.append([row[0],row[1], row[2], Sample_Type,Sample_Personalization,row[5],Shipping_Method,row[7],row[8],row[9],City_State_Zip,Special_Instructions])
else:
    check_list="全部樣版本: "
    for cp in check_point:
        cp1=df_Sample.loc[0,cp]
        check_list=check_list+dict2.get(cp1, "")+" "
    print(check_list)
    ws1.append((check_list,))
    ws1.append(df_Sample.columns.tolist())
    for index, row in df_Sample.iterrows():
        ws1.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9], row[10], row[11]])
    ws1.merge_cells("A1:L1")
    # print(Lot,Sample_Type)
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 5
ws1.column_dimensions['D'].width = 13
ws1.column_dimensions['E'].width = 23
ws1.column_dimensions['F'].width = 9
ws1.column_dimensions['G'].width = 16
ws1.column_dimensions['H'].width = 23
ws1.column_dimensions['I'].width = 26
ws1.column_dimensions['J'].width = 30
ws1.column_dimensions['K'].width = 30
ws1.column_dimensions['L'].width = 30

for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
    for cell_a in row:
        cell_a.border = border
        cell_a.alignment = alignment

wb.save("./test.xlsx")