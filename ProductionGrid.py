#update:2025-01-16 11:331aaass
from Share import Honour_Share
# from Share.Honour_Share import Py_Decrypto, update_ver,kill_process
import win32api
import win32con
from openpyxl.styles import Font, Color, PatternFill,Alignment,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook,load_workbook
import pandas as pd
import numpy as np
import re
import openpyxl
import clr
import xml.dom.minidom
import pymssql as sql
import os
import shutil
import configparser

import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QMessageBox
from PyQt5 import QtGui,QtCore

class MyApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # 设置窗口标题和大小
        self.setWindowTitle('简单的 PyQt5 示例')
        self.setGeometry(100, 100, 300, 200)

        # 创建一个垂直布局
        layout = QVBoxLayout()

        # 创建一个按钮并添加到布局中
        self.button = QPushButton('点击我', self)
        self.button.clicked.connect(self.showMessage)
        layout.addWidget(self.button)

        # 设置窗口的布局
        self.setLayout(layout)

    def showMessage(self):
        # 弹出一个消息框
        QMessageBox.information(self, '消息', '你点击了按钮！')

def ProductionGrid(wb,ws1,df,PresortClass,master_Component):
    # 创建一个边框对象
    font_e1 = Font(name='Calibri', size=11, bold=False, italic=False, color='000000')   #
    # font_e2 = Font(name='Calibri', size=11, bold=False, italic=False, color='FF00FF')
    font_e2=Font(name='Calibri', size=11, bold=False, italic=False, color='FF00FF')   #Font(color="FF00FF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    fill_1 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    list1="Lot","Job Number","Package Code","MailGrid","Production Site","Lettershop","Assembly Type","Assembly Option","Permit Number","Rate Class","Piece Type","OE Postage Type","Postage Strategy","Reply Classification","Reply Postage Type","Quantity"

    # group_PackageCode = df.groupby(["Package Code"])
    group_PackageCode = df.groupby(["pc_key"])
    # print("有多少个Group:",group_PackageCode.ngroups)
    for group_pc in group_PackageCode.groups:
        gp_pc = group_PackageCode.get_group((group_pc,))


        group_component=gp_pc.groupby(["Ord","Component"])
        gp_list1 = []
        gp_list1_b = []
        Folding_list_w = []
        stock_list_w = []

        for group1 in group_component.groups:
            row_pc = ws1.max_row
            # gp_Folding_list1 = []
            # gp_stock_list1 = []
            gp1 = group_component.get_group(group1)
            # print(group1)

            #Check Folding_desc
            group_Folding=gp1.groupby(["Folding_desc"])
            groups_check1=group_Folding.ngroups
            if groups_check1>1:
                Folding_info=[]
                for group2 in group_Folding.groups:
                    gp2 = group_Folding.get_group(group2)
                    Folding_Lot_info="Lot "
                    Total_row=gp2.shape[0]
                    Counter=1
                    for index, row in gp2.iterrows():
                        if Counter==Total_row:
                            Folding_Lot_info=Folding_Lot_info+str(row["Lot"])
                        else:
                            Folding_Lot_info = Folding_Lot_info + str(row["Lot"]) + ","
                        Counter+=1
                        # print("info:",row["Ord"],row["Component"],row["Item Code"],row["Lot"],row["Folding_desc"],row["stock_desc"])


                    Folding_info1=Folding_Lot_info+":"+group2+"\n"
                    # print("Folding_info1:",Folding_info1)
                    Folding_info.append(Folding_info1)
                # print("Folding_info:", Folding_info)
            else:
                Folding_info = list(group_Folding.groups.keys())
                # print("Folding_info_0:", Folding_info)

            #Check stock_desc
            group_stock=gp1.groupby(["stock_desc"])
            groups_check_b=group_stock.ngroups
            if groups_check_b>1:
                stock_info=[]
                for group2_b in group_stock.groups:
                    gp2_b = group_stock.get_group(group2_b)
                    stock_Lot_info="Lot "
                    Total_row_b = gp2_b.shape[0]
                    Counter_b = 1
                    for index_b, row_b in gp2_b.iterrows():
                        if Counter_b == Total_row_b:
                            stock_Lot_info = stock_Lot_info + str(row_b["Lot"])
                        else:
                            stock_Lot_info = stock_Lot_info+str(row_b["Lot"])+","
                        Counter_b += 1
                    # if stock_Lot_info[0]==',':
                    #     stock_Lot_info=stock_Lot_info[1:]
                    # if stock_Lot_info[-1]==',':
                    #     stock_Lot_info=stock_Lot_info[:-1]
                    stock_info1=stock_Lot_info+":"+group2_b+"\n"
                    stock_info.append(stock_info1)
                # print("stock_info:", stock_info)
            else:
                stock_info = list(group_stock.groups.keys())
                # print("stock_info_0:", stock_info)



            # print("大組:", group1, "小組數量:", groups_check1,"info:",Folding_info)
            gp_list1.append(group1)
            gp_list1_b.append(group1)
            t1=("\n".join(Folding_info),)
            t1_b = ("\n".join(stock_info),)
            # gp_Folding_list1.append(group1+t1)
            # gp_stock_list1.append(group1 + t1_b)
            # print(gp_Folding_list1)
            # print(gp_stock_list1)
            Folding_list_w.append(group1+t1)
            stock_list_w.append(group1 + t1_b)


        # wb = Workbook()
        # ws1 = wb.active

        # ws3["A1"] = "test"


        columns1 = ["Ord", "Component"]
        component_all = pd.DataFrame(gp_list1, columns=columns1)

        columns2 = ["Ord", "Component","Folding"]
        Folding_list = pd.DataFrame(Folding_list_w, columns=columns2)

        columns3 = ["Ord", "Component", "Stock Type"]
        stock_list = pd.DataFrame(stock_list_w, columns=columns3)

        component_all_add1 = pd.merge(component_all, master_Component, on=["Component"], how='left')  # Head
        component_all_add2 = pd.merge(component_all_add1, Folding_list, on=["Ord","Component"], how='left')  # Head
        component_all_add3 = pd.merge(component_all_add2, stock_list, on=["Ord", "Component"], how='left')  # Head
        component_all_add3["mark"] = component_all_add3["Stock Type"] + "\n" + component_all_add3["Folding"]

        #第一列更新數據
        len_list1=len(list1)
        for row,l in enumerate(list1):
            range_cell='A'+str(row+row_pc).strip()+":"+'B'+str(row+row_pc).strip()
            # print("Merge: ",range_cell)
            cell_v = 'A' + str(row+row_pc).strip()
            ws1.merge_cells(range_cell)
            ws1[cell_v]=l

            # test=ws1[cell_v]
            # print(ws1[cell_v])

        ws1.append(["Ord", "Component"])
        max_row1 = ws1.max_row
        # cell_m1="A" + str(ws1.max_row).strip()

        # ws1["A" + str(ws1.max_row).strip()].font=font_e2
        alignment = Alignment(vertical='center', wrap_text=True)    #horizontal='center'
        for index, row in component_all_add3.iterrows():
            Ord = row["Ord"]
            Component = row["Component"]
            desc = row["desc"]
            mark = row["mark"]
            if desc is np.nan:
                desc=""
            # cell_2A = 'A' + str(index+row_pc + 1 + len_list1).strip()

            cell_2A = 'A' + str(index+row_pc + 1 + len_list1).strip()
            cell_2B = 'B' + str(index+row_pc + 1 + len_list1).strip()
            ws1[cell_2A] = Ord
            ws1[cell_2A].alignment = alignment
            ws1[cell_2B] = Component+"\n"+desc
            ws1[cell_2B] .alignment = alignment
            # ws1[cell_2A] = mark
            # ws1[cell_2A].alignment = alignment
            # ws1[cell_2A].font=font_e2
        #第一列更新數據 end

        alignment2 = Alignment(vertical='center', horizontal='left')
        group_Lot = gp_pc.groupby(["Package Code","Lot"])

        for col1,group2 in enumerate(group_Lot.groups):
            gp2 = group_Lot.get_group(group2)
            # print(group2)
            Lot_add_PresortClass = pd.merge(gp2, PresortClass, on=["Rate Class","Piece Type"], how='left')  #Head
            # Lot_add_PresortClass = Lot_add_PresortClass.assign(Piece_Type="")

            Lot_add_PresortClass["Piece_Type"] = Lot_add_PresortClass["Piece Type"]+"\n限重: "+Lot_add_PresortClass["weight"].astype(str)+"g\n限厚: "+Lot_add_PresortClass["thickness"].astype(str)+"mm"
            Lot_Head=Lot_add_PresortClass[["Lot","Job Number","Package Code","MailGrid","Production Site","Lettershop","Assembly Type","Assembly Option","Permit Number","Rate Class","Piece_Type","OE Postage Type","Postage Strategy","Reply Classification","Reply Postage Type","Quantity"]]
            col_head=col1+3
            column_letter = get_column_letter(col_head)
            ws1.column_dimensions[column_letter].width = 16.5
            Head1 = Lot_Head.iloc[0]
            lot_component_count = gp2.shape[0]
            for row_head,value_head in enumerate(Head1.values):
                # print("value: ",value_head,"type:",type(value_head))
                if pd.isna(value_head):
                    value_head=""
                if type(value_head)==np.float64 or type(value_head)==np.int64:
                    # cell_a=column_letter+str(row_head+1).strip()
                    # ws3[cell_a].alignment = alignment2
                    value_head=str(value_head).strip()
                if row_head==14:    #Reply Postage Type 栏
                    value_head=value_head+"\n輔料件數(以貨號計):\n入外封件數(例: NP+BK為一件):\n圓形貼紙件數(如有):"


                ws1.cell(row=row_head+row_pc,column=col_head, value=value_head)

            # print("gp2:", gp2.shape[0])
            ws1.cell(row=row_head+row_pc + 1, column=col_head, value="Item Code("+str(lot_component_count).strip()+")")

            Lot_Component_all = pd.merge(component_all, gp2, on=["Ord", "Component"], how='left')

            for index, components in Lot_Component_all.iterrows():
                Item_Code = components["Item Code"]
                Direction = components["Direction"]
                Perso = components["Perso"]
                Buy_Out = components["Buy Out"]
                if pd.isna(Item_Code):
                    Item_Code_w=""
                elif type(Item_Code)==float or type(Item_Code)==int:
                    Item_Code_w = str(int(Item_Code))
                else:
                    Item_Code_w=Item_Code

                if Direction=="Front":
                    Direction_w="(+)"
                elif Direction=="Back":
                    Direction_w = "(-)"
                else:
                    Direction_w = ""

                if Perso=="✓":
                    Perso_w="(P)"
                else:
                    Perso_w=""

                if Buy_Out=="✓":
                    Buy_Out_w="(BO)"
                else:
                    Buy_Out_w=""
                Item_Code_w=Item_Code_w+Direction_w+Perso_w+Buy_Out_w
                ws1.cell(row=row_pc+17 + index, column=col_head, value=Item_Code_w)

        #更新摺法/用紙等信息
        max_col1 = col_head
        # max_col1 = ws1.max_column
        alignment = Alignment(vertical='center', wrap_text=True)    #horizontal='center'
        for index, row in component_all_add3.iterrows():
            Ord = row["Ord"]
            # Component = row["Component"]
            # desc = row["desc"]
            mark = row["mark"]

            ws1.cell(row=max_row1 + 1 + index, column=max_col1+1, value=mark)
            # ws1.cell(row=max_row1 + 1 + index, column=max_col1 + 2, value=Ord)
            ws1.cell(row=max_row1 + 1 + index, column=max_col1 + 1).font=font_e2
            ws1.cell(row=max_row1 + 1 + index, column=max_col1 + 1).alignment = alignment
        #更新摺法/用紙等信息 end

        for row in ws1.iter_rows(min_row=row_pc+9, max_row=row_pc+14, min_col=3, max_col=ws1.max_column):
            for cell in row:
                cell.alignment = alignment
        # ap_line=[]
        # print("Packagecode Group count:",group_Lot.ngroups)
        # for ap in range(group_Lot.ngroups+3):
        #     ap_line.append("---")
        ws1.append([""])
        row_pc2 = ws1.max_row
        row_mark = ws1[row_pc2]
        for cell in row_mark:
            if cell.column<(group_Lot.ngroups+4):
                cell.fill = fill_1
                cell.value = "***"

        ws1.append(["", "", ""])


    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 22
    # ws1.column_dimensions['C'].width = 22
    # 遍历工作表中的所有单元格并应用边框

    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell_a in row:
            cell_a.border = border
            # print(cell_a.font.color.value)
            if cell_a.font.color.value==1:
                cell_a.font = font_e1


    # ws1.delete_rows(1)
    return ws1

def BlankWeight_Excel(innsight_job, client, blankWeight,master_Component,PresortClass,PackageCode_d,ws2,ws3,ws4):

    """
    :param innsight_job:
    :param client:
    :param blankWeight: pandas datafram mg6
    :param PresortClass: Master table:PresortClass
    :param PackageCode_d: Master table:PackageCode
    :param wb: 工作簿
    :param ws2: 工作表: blankWeight
    :return:
    """
    def head_styles(ws, row_start,row_x):
        # fx3 = "A" + row_x.__str__()
        # hx3 = "I" + row_x.__str__()
        fx3 = "A" + row_start.__str__()
        hx3 = "I" + row_x.__str__()

        cell_ranges = ws[fx3:hx3]
        font_c1 = Font(name='SimSun', size=12, bold=True, italic=False, color='FF0000')

        fill_1 = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        for i_row in cell_ranges:
            for i_cell in i_row:
                i_cell.fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
                # i_cell.font = font_e1

    # master_file = './master/master.xlsx'
    # Template_file = './master/Template.xlsx'
    # dst_file = './'+innsight_job+".xlsx"
    # shutil.copy(Template_file, dst_file)

    # PresortClass = pd.read_excel(master_file, sheet_name="PresortClass")
    # PackageCode_d = pd.read_excel(master_file, sheet_name="PackageCode")

    # wb = load_workbook(dst_file)
    # 获取活动的工作表
    # ws1 = wb.active
    # ws1 = wb["ProductionGrid"]
    # ws2=wb["BlankWeight"]
    # ws1.title = "生產表"
    # ws1=wb.create_sheet("生產表")
    # ws2=wb.create_sheet("白樣資料")

    # ws = wb.active
    # # 给工作表命名（可选）
    # ws.title = "白樣資料"
    #list1_suffx = ['OE', 'CH', 'OB', 'OS', 'SB', 'SL', 'SM', 'XB']
    list1_suffx = ['OE - Paper','OE - Special', 'OE - Chipboard','OE - Japanese Style', 'Outer Box', 'Sandwich Bags - Reusable', 'Compact Disc Sleeve', 'Self Mailer','Outer Polybag']

    list2_suffx = ['RE', 'RC', 'LRC']
    list3_suffx = ['ME', 'CD']
    font_e1 = Font(name='Calibri', size=12, bold=False, italic=False, color='000000')
    # 创建一个边框对象
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 22
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['G'].width = 13
    ws2.column_dimensions['H'].width = 11
    ws2.column_dimensions['I'].width = 12

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 28
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 14
    ws3.column_dimensions['F'].width = 24
    ws3.column_dimensions['G'].width = 14
    ws3.column_dimensions['H'].width = 20

    ws3.append(["工單", "版本", "檢查項目", "檢查結果", "檢查貨號", "檢查尺寸", "異常貨號", "異常尺寸"])

    blankWeight2 = pd.merge(blankWeight, master_Component, left_on=["Component_x"], right_on=["Component"], how='left')  # on 字段不同处理方式

    bw1=blankWeight2.groupby("package_key")

    group_num=1
    ws4.column_dimensions['A'].width = 12
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 48
    ws4.column_dimensions['D'].width = 45
    ws4.column_dimensions['E'].width = 16
    ws4.column_dimensions['F'].width = 18
    ws4.append(["工單","版本","重量/厚度","郵資","外封/盒名稱","尺寸"])
    for group in bw1.groups:
        group_num=group_num+1
        # print(group)
        gp=bw1.get_group(group)
        # Job_number=""
        if innsight_job[0:2] == '40':
            HonourJob = "1" + innsight_job.replace("-", "")[2:8]
        elif innsight_job[0:2] == '41':
            HonourJob = "2" + innsight_job.replace("-", "")[2:8]


        component_qty=str(gp.iloc[0]["component_qty_y"])
        PackageCode = gp.iloc[0]["Package Code"]
        # package_key = gp.iloc[0]["package_key"]
        # Lot_list = gp.iloc[0]["Lot_list"]
        # Lot_list=str(group).split('-')[2]
        Lot_list2 = str(group).split('-')[2].strip()
        if Lot_list2[-1]==';':
            Lot_list2=Lot_list2[:-1]
        pc_key = gp.iloc[0]["pc_key"]
        # print("pc_key:",pc_key)
        PieceType=gp.iloc[0]["Piece Type"]
        RateClass = gp.iloc[0]["Rate Class"]
        # PresortClass
        search_PresortClass=PresortClass[(PresortClass["Piece Type"] == PieceType) & (PresortClass["Rate Class"] == RateClass)]
        if search_PresortClass.shape[0] != 0:
            PC_weight = search_PresortClass.iloc[0]["weight"]
            PC_thickness = search_PresortClass.iloc[0]["thickness"]
        else:
            PC_weight=0
            PC_thickness=0

        #PackageCode
        search_PackageCode=PackageCode_d[(PackageCode_d["pk_key"] == pc_key)]
        if search_PackageCode.shape[0] != 0:
            PC_jobnum = search_PackageCode.iloc[0]["Jobnum"]
            if str(PC_jobnum).find(',')>=0:
                joblist=PC_jobnum.split(',')
                joblist.sort(reverse=True)
                # ref_jobnum =",".join(joblist[:2])
                ref_jobnum = ""
                for refjob in joblist[:2]:
                    if refjob[0:2] == '40':
                        HonourJob2 = "1" + refjob[2:]
                    elif refjob[0:2] == '41':
                        HonourJob2 = "2" + refjob[2:]
                    ref_jobnum=ref_jobnum+HonourJob2+","

                # print("ref:",ref_jobnum)
            else:
                if PC_jobnum[0:2] == '40':
                    HonourJob3 = "1" + PC_jobnum[2:]
                elif PC_jobnum[0:2] == '41':
                    HonourJob3 = "2" + PC_jobnum[2:]
                ref_jobnum=HonourJob3

        else:
            ref_jobnum=""

        if client is None:
            ws2.append(["客戶", ""])
        else:
            ws2.append(["客戶",client[1]])
        row_start = ws2.max_row
        ws2.append(["工單", innsight_job,HonourJob,"","","","總件數",0,"件"])
        # ws2.append(["版本", Lot_list2.strip()[:-1],"","","","","總重量",0,"克"])
        ws2.append(["版本", Lot_list2, "", "", "", "", "白樣重量", 0, "克"])
        row_weight = ws2.max_row
        ws2.append(["Package Code", PackageCode,"","","","","白樣厚度",0,"mm"])
        ws2.append(["參考舊單", ref_jobnum, "","","","","限制重量",PC_weight,"克"])
        ws2.append(["郵資", PieceType,RateClass,"","","","限制厚度",PC_thickness,"mm"])
        ws2.append(["輔料("+component_qty+")","貨號", "平張尺寸","用紙","紙重","件數","個重","重量","備注"])
        row_x = ws2.max_row
        head_styles(ws2, row_start,row_x)
        # head_styles(ws2, row_x - 1)
        # head_styles(ws2, row_x - 3)
        # head_styles(ws2, row_x - 5)
        # Total_weight=0

        for index, row in gp.iterrows():
            # row_c = row["Component"]
            row_detail = ws2.max_row
            #weight=row["weight"] * row["times"]
            weight=f'=F{str(row_detail+1).strip()}*G{str(row_detail+1).strip()}'
            # Total_weight=Total_weight+weight
            times_PrintOption=str(row["Print Option"]).count('/')
            if str(row["Print Option"]).count('/')>1:
                times=row["times"]*times_PrintOption
            else:
                times = row["times"]
            ws2.append([row["Component_x"], row["Item Code"], row["Flat Size"], row["Stock Type"], row["Stock Option"], times, row["weight"], weight, row["templateCode"]])

        row_y = ws2.max_row
        ws2["H" + str(row_weight-1).strip()] = f'=SUM(F{str(row_x + 1).strip()}:F{str(row_y).strip()})'
        ws2["H"+str(row_weight).strip()] = f'=ROUND(SUM(H{str(row_x+1).strip()}:H{str(row_y).strip()}),2)'
        #字体加颜色 白样重量, 厚度

        # fx3 = "G" + row_weight.__str__()
        # hx3 = "I" + str(row_weight+1).strip()
        # cell_ranges2 = ws2[fx3:hx3]
        # font_c1 = Font(name='SimSun', size=14, bold=True, italic=False, color='FF0000')
        # # font_e1 = Font(name='Arial', size=12, bold=True, italic=False, color='000000')
        # for i_row2 in cell_ranges2:
        #     for i_cell2 in i_row2:
        #         i_cell2.font = font_c1
        # 字体加颜色 end
        #ddata 重量厚度信息
        weight_1=str(row_weight - 1).strip()
        weight_0=str(row_weight).strip()
        weight_add1 = str(row_weight + 1).strip()
        range_cell1='D'+str(row_start).strip()+":"+'I'+str(row_start).strip()
        ws2.merge_cells(range_cell1)
        #ws2["D" + str(row_start).strip()] = f'=B{weight_1}&B{weight_0}&"重量為:"&H{weight_0}&I{weight_0}&"("&ROUND(H{weight_0}*0.035274,3)&"oz); 厚度為: "&H{weight_add1}&I{weight_add1}&"("&ROUND(H{weight_add1}*0.0393701,3)&""")"'
        ws2["D" + str(row_start).strip()] = f'="重量為:"&H{weight_0}&I{weight_0}&"("&ROUND(H{weight_0}*0.035274,3)&"oz); 厚度為: "&H{weight_add1}&I{weight_add1}&"("&ROUND(H{weight_add1}*0.0393701,3)&""")"'

        ws4["A" + group_num.__str__()] = f'=BlankWeight!B{weight_1}'
        ws4["B" + group_num.__str__()] = f'=BlankWeight!B{weight_0}'
        ws4["D" + group_num.__str__()] = f'=BlankWeight!B{str(row_weight + 3).strip()}&" "&BlankWeight!C{str(row_weight + 3).strip()}'
        ws4["C" + group_num.__str__()] = f'=BlankWeight!D{str(row_start).strip()}'


        #=B2&B3&"重量為:"&H3&"("&ROUND(H3*0.035274,3)&"oz); 厚度為: "&H5&I5&"("&ROUND(H5*0.0393701,3)&""")"
        # ddata 重量厚度信息 end
        range_cell2='C'+str(row_start+5).strip()+":"+'D'+str(row_start+5).strip()   #邮资合并单元格
        range_cell3 = 'B' + str(row_start + 4).strip() + ":" + 'D' + str(row_start+4).strip()   #參考舊單合并单元格
        range_cell4 = 'B' + str(row_start + 3).strip() + ":" + 'C' + str(row_start + 3).strip()  # Package Code合并单元格
        range_cell5 = 'B' + str(row_start + 2).strip() + ":" + 'D' + str(row_start + 2).strip()  # 版本合并单元格
        range_cell6 = 'B' + str(row_start).strip() + ":" + 'C' + str(row_start).strip()  # 客户合并单元格
        ws2.merge_cells(range_cell2)
        ws2.merge_cells(range_cell3)

        #字体加颜色 白样重量, 厚度

        fx3 = 'B' + str(row_start + 4).strip()
        hx3 = 'D' + str(row_start+4).strip()
        cell_ranges2 = ws2[fx3:hx3]
        for i_row2 in cell_ranges2:
            for i_cell2 in i_row2:
                i_cell2.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # 字体加颜色 end

        ws2.merge_cells(range_cell4)
        ws2.merge_cells(range_cell5)
        ws2.merge_cells(range_cell6)
        ws2.append(["",""])

        # ws3 工作表(MatchSize) 处理
        # ws3.append(["工單", innsight_job,  "版本", Lot_list.strip()[:-1],"Package Code", PackageCode])

        check1=gp[gp["Component"].isin(list1_suffx)]
        if check1.shape[0]!=0:
            max_long_index = gp['long'].idxmax()
            max_width_index = gp['width'].idxmax()
            row_long=gp.loc[max_long_index]
            row_width = gp.loc[max_width_index]
            check_suffix1=row_long["Component"]
            check_suffix2 = row_width["Component"]
            if check_suffix1 in list1_suffx:
                ws4["E" + group_num.__str__()] = row_long["Item Code"]
                ws4["F" + group_num.__str__()] = row_long["Flat Size"]
            if check_suffix1==check_suffix2 and check_suffix1 in list1_suffx:
                # print("外封高度寬度是最大的: Item Code: ", row_long["Item Code"], row_long["Finished Size"])
                pass
                # ws3.append([innsight_job,Lot_list2,"封/盒高度寬度是否最大", "OK", row_long["Item Code"],row_long["Finished Size"]])
            else:
                Head1 = check1.iloc[0]
                add1=Head1["Item Code"]+Head1["Finished Size"]
                # print("外封高度寬度有問題:  ", add1, row_long["Item Code"], row_long["Finished Size"], row_width["Item Code"],row_width["Finished Size"])
                if check_suffix1 in list1_suffx:
                    ws3.append([innsight_job, Lot_list2, "封/盒高度寬度是否最大", "异常", Head1["Item Code"], Head1["Finished Size"], row_width["Item Code"], row_width["Finished Size"]])
                else:
                    ws3.append([innsight_job, Lot_list2, "封/盒高度寬度是否最大", "异常", Head1["Item Code"],Head1["Finished Size"],row_long["Item Code"], row_long["Finished Size"]])

        check2 = gp[gp["Suffix_All"].isin(list2_suffx)]
        if check2.shape[0]>1:
            max_long_index = check2['long'].idxmax()
            max_width_index = check2['width'].idxmax()
            row_long=check2.loc[max_long_index]
            row_width = check2.loc[max_width_index]
            check_suffix1=row_long["Suffix_All"]
            check_suffix2 = row_width["Suffix_All"]

            if check_suffix1==check_suffix2 and check_suffix1 =="RE":
                # print("回郵封高度寬度是最大的: Item Code: ", row_long["Item Code"], row_long["Finished Size"])
                pass
                # ws3.append([innsight_job,Lot_list2,"回郵封高度寬度是最大", "OK", row_long["Item Code"],row_long["Finished Size"]])
            else:
                Head1 = check2[check2["Suffix_All"]=='RE']
                # add1=Head1["Item Code"]+Head1["Finished Size"]
                Head2 = Head1.iloc[0]
                # print("回郵封高度寬度有問題:  ", Head2["Item Code"], Head2["Finished Size"], row_long["Item Code"], row_long["Finished Size"], row_width["Item Code"],row_width["Finished Size"])
                if check_suffix1=='RE':
                    ws3.append([innsight_job, Lot_list2, "回郵封高度寬度是否最大", "异常", Head2["Item Code"], Head2["Finished Size"], row_width["Item Code"], row_width["Finished Size"]])
                else:
                    ws3.append([innsight_job, Lot_list2, "回郵封高度寬度是否最大", "异常", Head2["Item Code"], Head2["Finished Size"], row_long["Item Code"], row_long["Finished Size"]])

        check3 = gp[gp["Suffix_All"].isin(list3_suffx)]
        if check3.shape[0]>1:
            max_long_index = check3['long'].idxmax()
            max_width_index = check3['width'].idxmax()
            row_long=check3.loc[max_long_index]
            row_width = check3.loc[max_width_index]
            check_suffix1=row_long["Suffix_All"]
            check_suffix2 = row_width["Suffix_All"]

            if check_suffix1==check_suffix2 and check_suffix1 =="ME":
                # print("西封高度寬度是最大的: Item Code: ", row_long["Item Code"], row_long["Finished Size"])
                pass
                # ws3.append([innsight_job,Lot_list2,"西封高度寬度是最大", "OK", row_long["Item Code"],row_long["Finished Size"]])
            else:
                Head1 = check3[check3["Suffix_All"]=='ME']
                # add1=Head1["Item Code"]+Head1["Finished Size"]
                Head2 = Head1.iloc[0]
                # print("西封高度寬度有問題:  ", Head2["Item Code"], Head2["Finished Size"], row_long["Item Code"], row_long["Finished Size"], row_width["Item Code"],row_width["Finished Size"])
                if check_suffix1=='ME':
                    ws3.append([innsight_job, Lot_list2, "西封高度寬度是否最大", "异常", Head2["Item Code"], Head2["Finished Size"], row_width["Item Code"], row_width["Finished Size"]])
                else:
                    ws3.append([innsight_job, Lot_list2, "西封高度寬度是否最大", "异常", Head2["Item Code"], Head2["Finished Size"], row_long["Item Code"], row_long["Finished Size"]])


        # ws3 工作表(MatchSize) 处理 end

    # 遍历工作表中的所有单元格并应用边框
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
        for cell in row:
            if cell.column==9 and str(cell.value).startswith("TP"):
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.border = border
            cell.font = font_e1

    ws3maxrow = ws3.max_row
    if ws3maxrow==1:
        ws3.append([innsight_job, "全部版本", "全部检查项目", "OK", "", "", "", ""])
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=ws3.max_column):
        for cell in row:
            cell.border = border
            cell.font = font_e1
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    # wb.save(dst_file)
    return ws2,ws3,ws4

# def Py_Decrypto():
#
#     dom = xml.dom.minidom.parse(r'C:\HonourProgram\Live\userCommon.xml')
#     root = dom.documentElement  # students结点
#     # print(root.getAttribute('userCommon'))
#     userIDSQL = root.getElementsByTagName('userIDSQL')
#     userPWDSQL=root.getElementsByTagName('userPWDSQL')
#
#     serverName = root.getElementsByTagName('serverName')[0].firstChild.data
#     databaseName = root.getElementsByTagName('databaseName')[0].firstChild.data
#
#     userID_Encrypt=userIDSQL[0].firstChild.data
#     userPWD_Encrypt=userPWDSQL[0].firstChild.data
#
#     clr.AddReference(r"C:\HonourProgram\Live\Appstore\Honour.dll")
#
#     # import the namespace and class
#
#     from Honour import SymmetricMethod
#
#     # create an object of the class
#
#     obj = SymmetricMethod()
#
#     v1=obj.Encrypto("beginer")
#     # print("加密: ",v1)
#     value = obj.Decrypto("ivTrAqqgiIMUD4RL31nRgA==")
#     # print("解密: ", value)
#
#     userID=obj.Decrypto(userID_Encrypt)
#     userPWD = obj.Decrypto(userPWD_Encrypt)
#     # print("解密: ", userID,userPWD)
#
#     return userID,userPWD,serverName,databaseName

def read_sql_fetchall(sql_str,userID, userPWD, serverName, databaseName):

    # conn = sql.connect(server='HonourSQL02\HonourSQL', user='beginer', password='@fly314', database='HMPSQL01')
    # conn = sql.connect(server='10.2.81.30', user='beginer', password='@fly314', database='HMPSQL01')
    conn = sql.connect(server=serverName, user=userID, password=userPWD, database=databaseName,tds_version="7.0")
    # stock_basic = conn.cursor(as_dict=True)

    stock_basic = conn.cursor()
    # stock_basic.execute("select  max(id) as num from [dbo].[mv_job] where jobver=%s",jobver)

    # sql_str=f"SELECT templateCode,suffix,widthmm1,widthmm2,UnitArea FROM [HMPSQL01].[dbo].[V_StdTemplate]  where suffix='{Suffix}' and ((widthmm1={m_long} and widthmm2={m_width}) or (widthmm1={m_width} and widthmm2={m_long})) order by suffix,widthmm1,widthmm2,UnitArea desc"
    stock_basic.execute(sql_str)

    #data=stock_basic.fetchone()
    all=stock_basic.fetchall()
    # print(all)
    # max_id = stock_basic.fetchone()
    # if max_id is None:
    #     max_id=0
    # jobver_id = jobver + "-" + str(max_id + 1).rjust(4, '0')
    # print(max_id)
    # values1=(max_id + 1,jobver,jobver_id,process_department)
    # stock_basic.execute("insert into  [dbo].[mv_job] ([id],[jobver],[jobver_id],[dept]) values(%d,%s,%s,%s)", values1)
    conn.commit()
    return all
    # print(jobver_id)

def read_sql_fetchone(sql_str,userID, userPWD, serverName, databaseName):


    # conn = sql.connect(server='HonourSQL02\HonourSQL', user='beginer', password='@fly314', database='HMPSQL01')
    # conn = sql.connect(server='10.2.81.30', user='beginer', password='@fly314', database='HMPSQL01')
    conn = sql.connect(server=serverName, user=userID, password=userPWD, database=databaseName,tds_version="7.0")
    # stock_basic = conn.cursor(as_dict=True)

    stock_basic = conn.cursor()
    # stock_basic.execute("select  max(id) as num from [dbo].[mv_job] where jobver=%s",jobver)

    # sql_str=f"SELECT templateCode,suffix,widthmm1,widthmm2,UnitArea FROM [HMPSQL01].[dbo].[V_StdTemplate]  where suffix='{Suffix}' and ((widthmm1={m_long} and widthmm2={m_width}) or (widthmm1={m_width} and widthmm2={m_long})) order by suffix,widthmm1,widthmm2,UnitArea desc"
    stock_basic.execute(sql_str)

    data=stock_basic.fetchone()
    # all=stock_basic.fetchall()
    # print(all)
    # max_id = stock_basic.fetchone()
    # if max_id is None:
    #     max_id=0
    # jobver_id = jobver + "-" + str(max_id + 1).rjust(4, '0')
    # print(max_id)
    # values1=(max_id + 1,jobver,jobver_id,process_department)
    # stock_basic.execute("insert into  [dbo].[mv_job] ([id],[jobver],[jobver_id],[dept]) values(%d,%s,%s,%s)", values1)
    conn.commit()
    return data
    # print(jobver_id)

def get_weight(size,gsm):

    re1 = re.compile(r'\s?([0-9,\.]{1,8})[a-z,",\s]{0,15}\sx', re.IGNORECASE)    #取size
    re2 = re.compile(r'x\s?([0-9,\.]{1,8})[a-z,"]{0,10}\s?', re.IGNORECASE)      #取size
    re3 = re.compile(r'\s?[0-9,\.]{1,8}([a-z,",\s]{0,15}\s)x', re.IGNORECASE)    #單位
    re4 = re.compile(r'\s?([0-9,\.]{1,8})\s{0,1}gsm', re.IGNORECASE)    #取gsm
    s1max = re1.findall(size)
    s2max = re2.findall(size)
    s3 = re3.findall(size)
    s4=re4.findall(gsm)
    # print(size)
    # print("gsm: ",gsm, s4)
    if s1max and s2max:
        if gsm=="House PSL Stock" or gsm=="House PSL Stock-MATTE":
            gsm_float=190
        elif s4:
            gsm_float=float(s4[0])
        else:
            gsm_float = 0
        first_Long = max(float(s1max[0]), float(s2max[0]))
        first_width = min(float(s1max[0]), float(s2max[0]))
        if str(s3[0]).strip().lower()=='cm':
            m_long=first_Long*10
            m_width=first_width*10
        elif str(s3[0]).strip().lower()=='mm':
            m_long=first_Long
            m_width=first_width
        elif str(s3[0]).strip().lower()=='' or str(s3[0]).strip().lower()=='"':
            m_long=first_Long*25.4    #寸轉為毫米mm=25.4
            m_width=first_width*25.4    #寸轉為毫米mm=25.4
        # print(size,gsm, m_long,m_width,"=",m_long*m_width*gsm_float)
        return max(round(m_long,0),round(m_width,0)),min(round(m_long,0),round(m_width,0)),gsm_float


    else:
        return 0,0,0

def main(PresortClass,PackageCode_d,master_Component,master_TemplateSuffix,master_Folding,master_stockType):
    print("生產表處理中...")
    # if not os.path.exists(r'C:\HonourProgram\Live\userCommon.xml'):
    #     shutil.copyfile('./Source/userCommon.xml',r'C:\HonourProgram\Live\userCommon.xml')
    #     print("update-UserCommon")
    # if not os.path.exists(r"C:\HonourProgram\Live\Appstore\Honour.dll"):
    #     shutil.copyfile('./Source/Honour.dll',r"C:\HonourProgram\Live\Appstore\Honour.dll")
    #     print("update-Honour.dll")

    userID, userPWD, serverName, databaseName =Honour_Share.Py_Decrypto(os.path.abspath('./Source/userCommon.xml'),os.path.abspath('./Source/Honour.dll'))
    excel_filepath=os.environ['USERPROFILE'] + '\\Downloads\\'

    # config = configparser.ConfigParser()
    # config.read("./config.ini", "utf-8-sig")  # utf-8-sig  & UTF-8
    # Master_Path = config['DEFAULT']['Master_Path']
    # master_file = Master_Path+'/master.xlsx'
    file_path1=excel_filepath+"data (1).xlsx"
    file_path2=excel_filepath+"data.xlsx"
    if not os.path.exists("./excel_file"):
        os.mkdir('./excel_file')
    if os.path.exists(file_path1) and os.path.exists(file_path1):
        df1 = pd.read_excel(file_path1, skiprows=2)
        df2 = pd.read_excel(file_path2, skiprows=2)
        # PresortClass = pd.read_excel(master_file, sheet_name="PresortClass")
        # PackageCode_d = pd.read_excel(master_file, sheet_name="PackageCode")
        # master_Component = pd.read_excel(master_file, sheet_name="Component")
        # master_TemplateSuffix = pd.read_excel(master_file, sheet_name="TemplateSuffix")
        # master_Folding = pd.read_excel(master_file, sheet_name="Folding")
        # master_stockType = pd.read_excel(master_file, sheet_name="stockType")

        wb = Workbook()

        ws3 = wb.active
        ws3.title = "MatchSize"
        ws1 = wb.create_sheet("ProductionGrid")
        ws2 = wb.create_sheet("BlankWeight")
        ws4 = wb.create_sheet("BlankWeightForDdata")


        # ws1 = wb.active
        # ws1.title = "ProductionGrid"
        # ws2 = wb.create_sheet("BlankWeight")
        # ws3 = wb.create_sheet("MatchSize")

        l0=df1.loc[0]
        cl1=df1.columns.values[0]
        cl2=df1.columns.values[0]
        if cl1=="Lettershop":
            fold_stock=df1
            insert=df2
        elif cl1=="Ord":
            insert = df1
            fold_stock = df2

        fold_stock = fold_stock.assign(Lot_ItemCode_qty=1)
        Check_pivot = pd.pivot_table(fold_stock, index=["Lot","Component","Item Code"], aggfunc={"Lot_ItemCode_qty": 'sum'})
        Check_pivot2=Check_pivot[Check_pivot["Lot_ItemCode_qty"]>1]
        Check_data_File = True
    else:
        Check_data_File = False

    if Check_data_File == False:
        win32api.MessageBox(0, "下载文件夹里有data.xlsx 或都 data(1).xlsx 不存在! 请检查!", "錯誤提示!", win32con.MB_OK)
    elif Check_pivot2.shape[0]>0:     #fold_stock 表Lot+Item Code不能有重复, 如有需要提示.
        err1 = ""
        for cp in Check_pivot2.itertuples():
            print(cp[0][0],cp[0][2],cp[1])
            err1=err1+"Lot " +str(cp[0][0]).strip() + " - Item Code: " + cp[0][2] + "重复. 重复次数 "+str(cp[1]).strip() +"; \n"
        win32api.MessageBox(0, "拆法用纸Excel data 表数据异常:"+"\n" + err1 +"\n请检查Excel data再处理.", "錯誤提示!", win32con.MB_OK)

    else:
        innsight_job=insert.values[0][8]
        campaignid=innsight_job[:7]
        phasenum=innsight_job[-1]
        dst_file = './' + innsight_job + ".xlsx"

        sql_job_str =f"SELECT  [InnJob],[str_abbreviation] FROM [HMPSQL01].[dbo].[V_Innsight_ForDataDB] where campaignid={campaignid} and phasenum={phasenum}"
        client=read_sql_fetchone(sql_job_str, userID, userPWD,serverName,databaseName)
        # print(client[1])

        # insert["Job Number","Lot","Package Code","Item Code","Rate Class","Piece Type"]
        mg=pd.merge(insert,fold_stock,on=["Lot","Item Code"])

        # mg_p1 = pd.merge(mg, master_Folding, on=["Folding"], how='left')
        # mg_p2 = pd.merge(mg_p1, master_stockType, on=["Stock Type"], how='left')
        # mg_p2.rename(columns={"Component_x":"Component","Direction_x":"Direction","Perso_x":"Perso","Lettershop_x":"Lettershop"}, inplace=True)
        # mg_p2 = mg_p2.assign(gsm=0)
        # mg_p2 = mg_p2.fillna("")
        # for index, row in mg_p2.iterrows():
        #     Stock_Option = row["Stock Option"]
        #     m_long, m_width, gsm_float = get_weight("2 X 4", Stock_Option)
        #     mg_p2.at[index, 'gsm'] = gsm_float
        #
        # mg_p2.loc[mg_p2["stock_desc"] != '', "stock_desc"] = mg_p2["gsm"].astype(str)+"克"+mg_p2["stock_desc"]


        blank_weight=mg[["Ord","Job Number","Lot","Package Code", "Piece Type","Rate Class","Component_x","Item Code","Flat Size","Stock Type","Stock Option","Times Used","Finished Size","Folding","MailGrid","Production Site","Lettershop_x","Assembly Type","Assembly Option","Permit Number","OE Postage Type","Postage Strategy","Reply Classification","Reply Postage Type","Quantity","Direction_x","Perso_x","Buy Out","Print Option"]]
        # blank_weight["pc_key"]=[np.nan] * len(blank_weight)
        # blank_weight["pc_key"]=[""] * len(blank_weight)
        blank_weight = blank_weight.assign(pc_key=np.nan,package_key="", LotQty=1, component_qty=1, delete_lot="")

        #此方法在原表顯示每個分組的統計數.
        # group_pjl=blank_weight.groupby(["Package Code","Job Number","Lot"])
        # blank_weight['C_qty']=group_pjl['component_qty'].transform('sum')

        # blank_weight.sort_values(by=["Package Code","Job Number","Lot","Ord"], inplace=True)
        #pivot1=按"Package Code","Job Number","Lot"分組統計component數量.
        pivot1=pd.pivot_table(blank_weight,index=["Package Code","Job Number","Lot"],aggfunc={"component_qty":'sum'})
        #on作为连接键的字段，当左右两个表的列名相同时使用。如果不相同，需要用left_on和right_on来分别指定.
        mg2=pd.merge(blank_weight,pivot1, on=["Package Code","Job Number","Lot"])
        mg2["pc_key"] = mg2["Package Code"] +"-"+mg2["component_qty_y"].astype(str)

        #
        mg_p1 = pd.merge(mg2, master_Folding, on=["Folding"], how='left')
        mg_p2 = pd.merge(mg_p1, master_stockType, on=["Stock Type"], how='left')
        mg_p2.rename(columns={"Component_x":"Component","Direction_x":"Direction","Perso_x":"Perso","Lettershop_x":"Lettershop"}, inplace=True)
        mg_p2 = mg_p2.assign(gsm=0)
        mg_p2 = mg_p2.fillna("")
        for index, row in mg_p2.iterrows():
            Stock_Option = row["Stock Option"]
            m_long, m_width, gsm_float = get_weight("2 X 4", Stock_Option)
            mg_p2.at[index, 'gsm'] = gsm_float

        mg_p2.loc[mg_p2["stock_desc"] != '', "stock_desc"] = mg_p2["gsm"].astype(str)+"克"+mg_p2["stock_desc"]

        #



        pivot2=pd.pivot_table(mg2,index=["pc_key","Lot"],aggfunc={"LotQty":'sum'})
        pivot2["LotQty"]=1

        group_pckey=pivot2.groupby(["pc_key"])

        #new
        lot_str=""
        lot_list=[]
        for group1 in group_pckey.groups:
            gp1 = group_pckey.get_group((group1,))
            for index, row in gp1.iterrows():
                lot_str = lot_str + str(index[1]) + "; "
            lot_list.append([group1,lot_str])
            lot_str = ""
        #new end
    ################
        # lot_str=""
        # lot_list=[]
        # for name,group in group_pckey:
        #     # print("package Code: ",name)
        #
        #     for gp in group.index:
        #         # print("Lot: ",gp[1])
        #         lot_str=lot_str+str(gp[1]).strip()+"; "
        #     lot_list.append([name,lot_str])
        #     lot_str = ""
    #####################

        columns1=["pc_key","Lot_list"]
        packageCode_Lotlist=pd.DataFrame(lot_list,columns=columns1)

        pivot3=pd.pivot_table(pivot2,index=["pc_key"],aggfunc={"LotQty":'sum'})
        mg3=pd.merge(mg2,pivot3, on=["pc_key"])

        GroupBypackagecodeQty=pd.pivot_table(mg3,index=["pc_key","LotQty_y","Flat Size"],aggfunc={"component_qty_x":'sum'})

        PackageCodeList=[]
        for gpc1 in GroupBypackagecodeQty.items():
            gpc2=gpc1[1]
            for gpc in gpc2.items():
                PackageCodeList.append([gpc[0][0],gpc[0][1],gpc[0][2],gpc[1],"No"])

        columns=["pc_key","Lot_Qty","Flat_size","Component_Qty","mergeLot"]
        df_packagecode=pd.DataFrame(PackageCodeList,columns=columns)
        df_packagecode.loc[df_packagecode["Component_Qty"] % df_packagecode["Lot_Qty"]==0, "mergeLot"] = "Yes"

        df_packagecode2=df_packagecode[["pc_key","mergeLot"]]

        df_packagecode3=df_packagecode2[df_packagecode2['mergeLot']=="No"]
        df_packagecode4 = df_packagecode3.drop_duplicates(subset=['pc_key','mergeLot'])

        df_packagecode5=pd.merge(df_packagecode2,df_packagecode4, on=["pc_key"], how='left')
        df_packagecode5.loc[df_packagecode5["mergeLot_y"] !='No', "mergeLot_y"] = "Yes"
        df_packagecode6 = df_packagecode5.drop_duplicates(subset=['pc_key','mergeLot_y'])
        mg4=pd.merge(mg3,df_packagecode6, on=["pc_key"], how='left')

        mg5=pd.merge(mg4,packageCode_Lotlist, on=["pc_key"])
        # mg5 = mg5.assign(delete_lot="")
        mg5.loc[mg5["mergeLot_y"] =='No', "package_key"] =mg5["pc_key"]+"-Lot "+mg5["Lot"].astype(str)
        mg5.loc[mg5["mergeLot_y"] =='Yes', "package_key"] =mg5["pc_key"]+"-Lot "+mg5["Lot_list"]


        mg5.loc[mg5["mergeLot_y"] =='Yes', "delete_lot"] =mg5["Lot_list"].str.extract(r';(.*)')[0]
        # mg5.sort_values(by=["delete_lot"])
        df_deleteLot = mg5.drop_duplicates(subset=['delete_lot'])
        # 使用iterrows()迭代行, 可以用列名稱row["delete_lot"]

        delectLot_str=""
        for index, row in df_deleteLot.iterrows():
            row_c=row["delete_lot"]
            if len(row_c)>1:
                delectLot_str=delectLot_str+row_c
        delectLot_list=delectLot_str.strip()[:-1].split(';')
        if delectLot_list[0]!="":
            delectLot_list2 = [int(item.strip()) for item in delectLot_list]
            #移除列個delectLot_list2的Lot版本
            mg6=mg5[~mg5['Lot'].isin(delectLot_list2)]
        else:
            mg6=mg5
        mg6 = mg6.assign(long=0, width=0, gsm=0, times=1, weight=0.0,templateCode="")

        #
        sql_StdTemplate_str="SELECT templateCode,suffix,long,width,UnitArea FROM [HMPSQL01].[dbo].[V_StdTemplate]"
        data_StdTemplate=read_sql_fetchall(sql_StdTemplate_str,userID, userPWD,serverName,databaseName)
        columns_StdTemplate=["templateCode","suffix","long","width","UnitArea"]
        StdTemplate=pd.DataFrame(data_StdTemplate,columns=columns_StdTemplate)
        #
        mg7 = pd.merge(mg6, master_TemplateSuffix, on=["Component_x"], how='left')
        mg7 = mg7.fillna("")
        # re_suffx = re.compile('\s?[0-9]{5,6}([a-z]{1,10})[0-9]{0,2}', re.IGNORECASE)    #取
        list_suffx=['OE','CH','OB','OS','ME','RE','REC','SB','SL','SM','XB']
        for index, row in mg7.iterrows():
            Component=row["Component_x"]
            Suffix = row["Suffix"]
            Item_Code=row["Item Code"]
            Flat_Size=row["Flat Size"]
            Finished_Size = row["Finished Size"]
            Stock_Option=row["Stock Option"]
            Times_Used=row["Times Used"]
            m_long, m_width, gsm_float = get_weight(Flat_Size, Stock_Option)  # Flat size,gsm,times
            f_long, f_width, gsm_float = get_weight(Finished_Size, Stock_Option)  # Finished_Size,gsm,times
            # Suffix = re_suffx.findall(Item_Code)
            # if str(Component)=="Reply Envelope":
            #     Suffix = ["RE"]
            # if str(Component)=="Reply Envelope-Bangtail":
            #     Suffix = ["REC"]
            # if str(Component)=="Matching Envelope":
            #     Suffix = ["ME"]

            mg7.at[index, 'long'] = f_long
            mg7.at[index, 'width'] = f_width
            if len(Suffix)>1:
                # mg6.at[index, 'Suffix'] = Suffix[0]

                search_StdTemplate = StdTemplate[(StdTemplate["suffix"] == Suffix) & (StdTemplate["long"] == m_long) & (StdTemplate["width"] == m_width)]
                if search_StdTemplate.shape[0] != 0:
                    UnitArea = float(search_StdTemplate.iloc[0]["UnitArea"])
                    templateCode = search_StdTemplate.iloc[0]["templateCode"]
                    # print("找到",Suffix[0],m_long,m_width,UnitArea)
                else:
                    UnitArea = 0
                    templateCode=""

                mg7.at[index, 'weight'] = round(UnitArea * 0.00064516 * gsm_float, 2)
                mg7.at[index, 'templateCode'] = templateCode
            else:
                if str(Component).startswith("NP -") or str(Component).startswith("NP-"):   #NP 要乘多少張
                    com_sp=Component.index('-')+1
                    # test=str(Component[com_sp:])
                    if str(Component[com_sp:]).strip().isnumeric():
                        NP_times=int(Component[com_sp:])
                        # mg6.at[index, 'Times Used'] = Times_Used*NP_times
                        # mg6.at[index, 'weight'] = (m_long/1000)*(m_width/1000)*gsm_float*NP_times
                        mg7.at[index, 'weight'] = (m_long / 1000) * (m_width / 1000) * gsm_float
                        mg7.at[index, 'times'] = NP_times
                    else:
                        mg7.at[index, 'weight'] = (m_long / 1000) * (m_width / 1000) * gsm_float
                elif str(Component)=="Mailing Label":   #貼紙貼在OE里, 要撕走底的.撕走的不要計算.
                    mg7.at[index, 'weight'] = ((m_long / 1000) * (m_width / 1000) * gsm_float) /2
                elif not str(Item_Code).startswith("Q"):    #貨號不是Q開頭的就計算重量.
                    mg7.at[index, 'weight'] = (m_long/1000)*(m_width/1000) * gsm_float

        ws1=ProductionGrid(wb,ws1,mg_p2, PresortClass, master_Component)
        ws2,ws3,ws4=BlankWeight_Excel(innsight_job,client,mg7,master_Component,PresortClass,PackageCode_d,ws2,ws3,ws4)

        try:
            wb.save(dst_file)
            data1 = "./excel_file/" + innsight_job + "_" + os.path.basename(file_path1)
            data2 = "./excel_file/"+innsight_job+"_" + os.path.basename(file_path2)
            #测试可以不用移源档案
            if os.environ['USERNAME']!="ITProg02":
                shutil.move(file_path1, data1)
                shutil.move(file_path2, data2)
            # 测试可以不用移源档案 end
            print("處理后的文件: ", os.path.abspath(dst_file))
            # Excel文件的路径
            # file_path = 'path_to_your_excel_file.xlsx'
            # 构建打开Excel文件的命令
            # command = ['start', 'excel', os.path.abspath(dst_file)]  # 注意：在Windows上，'start'是打开新窗口的命令
            command = f'start excel "{os.path.abspath(dst_file)}"'

            # 执行命令
            os.system(command)
            # subprocess.run(command)
        except Exception as e:
            print(e.args[1])
            if e.args[1]=='Permission denied':
                win32api.MessageBox(0, "Excel 文件: " + innsight_job+"已打开, 不能保存. 请关闭再处理.","錯誤提示!",win32con.MB_OK)
class MyApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        # 设置窗口标题和大小
        self.setWindowTitle('Innsight PBI 数据生成生产表&白样等数据')
        self.setGeometry(5, 35, 350, 100)

        # 创建一个垂直布局
        layout = QVBoxLayout()

        # 创建一个按钮并添加到布局中
        self.button = QPushButton('Innsight PBI 数据生成生产表&白样等数据', self)

        self.button.setStyleSheet("color: red;")
        self.button.setFixedSize(350,50)
        font = QtGui.QFont()
        font.setFamily("Microsoft Sans Serif")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.button.setFont(font)

        self.button.clicked.connect(self.run_production)
        layout.addWidget(self.button)

        # 设置窗口的布局
        self.setLayout(layout)

    def run_production(self):
        # 弹出一个消息框
        # QMessageBox.information(self, '消息', '你点击了按钮！')
        self.button.setEnabled(False)
        main(PresortClass,PackageCode_d,master_Component,master_TemplateSuffix,master_Folding,master_stockType)
        self.button.setEnabled(True)


# def update_ver1():
#     # filevers 版本+1
#     with open("./ver1.txt", 'r', encoding='utf-8') as file:
#         file_contents = file.read()
#         file.seek(0)
#         for line in file:
#             line=line.strip()
#             if line.startswith('filevers='):
#                 old_filevers = line
#                 ver = line.split('=')[1]
#                 ver1=ver[1:-2].split(',')
#                 ver2 = [int(i) for i in ver1]
#                 if ver2[3]>999:
#                     ver2[3]=0
#                     ver2[2] = ver2[2] + 1
#                 if ver2[2]>999:
#                     ver2[2] = 0
#                     ver2[3] = 0
#                     ver2[1] = ver2[1] + 1
#                 if ver2[1]>999:
#                     ver2[1] = 0
#                     ver2[2] = 0
#                     ver2[3] = 0
#                     ver2[0] = ver2[0] + 1
#                 else:
#                     ver2[3] = ver2[3] + 1
#                 new_filevers = "filevers=" + tuple(ver2).__str__()+","
#                 print("文件版本:", new_filevers)
#
#     #进行替换操作
#     updated_contents = file_contents.replace(old_filevers, new_filevers)
#
#     # 将修改后的内容写回文件（可选：可以写到一个新文件）
#     with open("./ver1.txt", 'w', encoding='utf-8') as file:
#         file.write(updated_contents)

if __name__ == '__main__':
    # frozen=hasattr(sys, 'frozen')       #打包為EXE后, frozen=True.  .Py 文件是False.
    # # win32api.MessageBox(0, str(frozen), "錯誤提示!", win32con.MB_OK)
    # if not frozen:
    #     update_ver("./ver1.txt")
    Honour_Share.kill_process('PyApp_')
    Honour_Share.update_ver("./ver_ProductionGrid.txt")
    config = configparser.ConfigParser()
    config.read("./Source/config_Production.ini", "utf-8-sig")  # utf-8-sig  & UTF-8
    Master_Path = config['DEFAULT']['Master_Path']
    Show_UI = config['DEFAULT']['Show_UI']
    master_file = Master_Path+'/master.xlsx'
    PresortClass = pd.read_excel(master_file, sheet_name="PresortClass")
    PackageCode_d = pd.read_excel(master_file, sheet_name="PackageCode")
    master_Component = pd.read_excel(master_file, sheet_name="Component")
    master_TemplateSuffix = pd.read_excel(master_file, sheet_name="TemplateSuffix")
    master_Folding = pd.read_excel(master_file, sheet_name="Folding")
    master_stockType = pd.read_excel(master_file, sheet_name="stockType")
    if Show_UI == "No":
    # if os.environ['USERNAME'] == "ITProg02":
        main(PresortClass,PackageCode_d,master_Component,master_TemplateSuffix,master_Folding,master_stockType)
    else:
        app = QApplication(sys.argv)
        ex = MyApp()
        ex.show()
        sys.exit(app.exec_())