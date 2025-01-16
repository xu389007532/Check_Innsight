import numpy as np
import math
import xlwings
from xlwings import Book,Sheet
# import xlwings as xw
from openpyxl.styles import Font, Color, PatternFill,Alignment,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook,load_workbook
import os.path
from os.path import *
import win32com.client
import configparser
import pandas as pd
import re
import shutil
from copy import copy

class constant():
    master_sheet_no_cn="主檔無中文"
    gsm_na = "找不到克重!"

    # stockType_na = "主檔找不到此紙類!"
    # stockType_no_cn = "紙類主檔無中文!"
    # Folding_na = "主檔找不到此折法."
    # Folding_no_cn = "折法主檔無中文!"
    # perforation_na = "主檔找不到此啤法."
    # perforation_no_cn = "啤主檔無中文!!"
    # imagingOption_na = ""
    personalizationOption_na = ""

class master():

    def __init__(self,master_file,master_sheet):
        self.master = pd.read_excel(master_file, sheet_name=master_sheet)
        self.master_dict = copy(self.master)
        self.master_sheet_no_cn = "主檔無中文"
        self.gsm_na = "找不到克重!"
        for m in self.master:
            ms = self.master.get(m)
            # ms_notna = ms[ms.iloc[:, 1].notna()]  # 移去第1列為na 的數據. 是從0開始的.
            dict1 = {}
            for st in ms.itertuples():
                dict1[st[1]] = st[2]
                # print(st[0], st[1], st[2])
            self.master_dict[m] = dict1


    def get_dict(self,sheet,sheet_key,sheet_key_default="show find out info!"):
        """
        :param sheet:
        :param sheet_key:
        :param sheet_key_default: 默認找不到key 返回: sheet + ":" + str(sheet_key) + "在主檔找不到.", 否則就返回 sheet_key_default 內容
        :return:
        """
        if sheet_key:
            # print(self.master_dict.get(sheet, '').get(sheet_key, ""))
            # print("fanyi: ", sheet, sheet_key)# , sheet_value, type(sheet_value))
            if sheet_key_default=="show find out info!":
                sheet_key_default_value = sheet + ":" + str(sheet_key) + "在主檔找不到."
            else:
                sheet_key_default_value = sheet_key_default
            sheet_value=self.master_dict.get(sheet).get(sheet_key, sheet_key_default_value)

            if isinstance(sheet_value,float) or isinstance(sheet_value, int):
                if math.isnan(sheet_value):
                    sheet_value = sheet+":"+sheet_key+self.master_sheet_no_cn
            elif sheet_value==None:
                sheet_value = sheet+":"+sheet_key + self.master_sheet_no_cn
            # elif sheet=="Sample":
            #     sheet_value=sheet_value
            #     print(sheet_value)
            return sheet_value
        else:
            return ""

def read_lotus_Instruction(innsight_job):
    config = configparser.ConfigParser()
    config.read("./config.ini", "utf-8-sig")  # utf-8-sig  & UTF-8
    Lotus_server  = config['DEFAULT']['Lotus_server']
    s = win32com.client.Dispatch('Notes.NotesSession')
    db = s.GetDatabase(Lotus_server, r"PublicNSF\Instruct.nsf")
    view = db.GetView(r"31. Instruction\ 01. by Item #")
    dc = view.GetAllDocumentsByKey(innsight_job, True)
    HK_filename=""
    filename = ""
    for i in range(1,dc.count+1):
        doc = dc.GetNthDocument(i)
        # jl=doc.GetItemValue("ItemNumber")
        rtitem = doc.GetFirstItem("ClientInstruction")

        if rtitem is not None:
            for r in rtitem.EmbeddedObjects:
                print(r.Type, r.name)
                rname = r.Name
                file_type = os.path.splitext(rname)[-1].lower()
                if file_type in ['.xlsx','.xls'] and len(rname)<15:  # 如果是Excel檔并且長度小15位的, 定為HK放的檔案, 就取出來處理.
                    HK_filename = os.getcwd()+"/excel_file/HK_" + r.Name
                    filename=os.getcwd()+"/excel_file/DIFS_" + r.Name
                    print("HK data: ",HK_filename)
                    r.ExtractFile(HK_filename)
    return HK_filename,filename

# read_lotus_Instruction("test")
def get_Country(Cit_State_Zip):
    # 定义美国地址正则
    us_address_re = re.compile(
        r'(?P<city>[A-Za-z ]+), '
        r'(?P<state>[A-Z]{2})\s+'
        r'(?P<zipcode>\d{5}(-\d{4})?)$'
    )
    match_address = us_address_re.match(Cit_State_Zip)
    if match_address:
        # print(f"Valid Address: {match_address.group()}")
        # print(match_address.groupdict())
        return "美國"
    else:
        if Cit_State_Zip.endswith(", NL"):
            return "荷蘭"
        # elif Cit_State_Zip.endswith(", IT"):
        #     return "意大利"
        else:
            return ""
def sample(ws1,HK_job,sample_sheet,master):
    alignment = Alignment(vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # excel=pd.ExcelFile(HK_job)
    # Sample_check=False
    # for sn in excel.sheet_names:
    #     if str(sn).lower().startswith("sample"):
    #         Sample_check=True
    #         sample_sheet=sn

    # dict1 = master_Sample.to_dict()
    # keys = list(dict1["Sample_eng"].values())
    # values = list(dict1["Sample_cn"].values())
    # dict2 = {key: value for key, value in zip(keys, values)}

    df_Sample = pd.read_excel(HK_job, sheet_name=sample_sheet)
    df_Sample=df_Sample.dropna(axis=0,subset=['Quantity'])
    df_Sample = df_Sample.fillna("")
    check_point = ["Sample Type", "Sample Personalization", "Shipping Method", "City State Zip", "Special Instructions"]
    Sample_group = df_Sample.groupby(by=check_point)
    if Sample_group.groups.__len__() > 2:  # 有多個不同的sample翻譯
        ws1.append(df_Sample.columns.tolist())
        for index, row in df_Sample.iterrows():
            Lot = row["Lot"]
            Sample_Type_eng = row["Sample Type"]
            Sample_Personalization_eng = row["Sample Personalization"]
            Shipping_Method_eng = row["Shipping Method"]
            City_State_Zip_eng = row["City State Zip"]
            Special_Instructions_eng = row["Special Instructions"]

            # Sample_Type = Sample_Type_eng + '\n' + dict2.get(Sample_Type_eng, "")
            Sample_Type = Sample_Type_eng + '\n' + master.get_dict("Sample",Sample_Type_eng)
            # Sample_Personalization = Sample_Personalization_eng + '\n' + dict2.get(Sample_Personalization_eng, "")
            Sample_Personalization = Sample_Personalization_eng + '\n' + master.get_dict("Sample",Sample_Personalization_eng)
            if Shipping_Method_eng=="Container":
                if Country!="":
                    Country="("+get_Country(City_State_Zip_eng)+")"
            else:
                Country = ""
            # Shipping_Method = Shipping_Method_eng + '\n' + dict2.get(Shipping_Method_eng, "")+Country
            Shipping_Method = Shipping_Method_eng + '\n' + master.get_dict("Sample",Shipping_Method_eng) + Country

            # Special_Instructions = Special_Instructions_eng + '\n' + dict2.get(Special_Instructions_eng, "")
            Special_Instructions = Special_Instructions_eng + '\n' + master.get_dict("Sample",Special_Instructions_eng,"")
            ws1.append([row.iloc[0], row.iloc[1], row.iloc[2], Sample_Type, Sample_Personalization, row.iloc[5], Shipping_Method, row.iloc[7], row.iloc[8], row.iloc[9], row.iloc[10], Special_Instructions])
    else:
        check_list = "全部樣版: "
        for cp in check_point:
            cp1 = df_Sample.loc[0, cp]
            if cp == 'City State Zip':
                Country = "(" + get_Country(cp1) + ")"
                check_list = check_list + Country + " "

            check_list = check_list + master.get_dict("Sample",cp1,"") + " "
        print(check_list)
        ws1.append((check_list,))
        ws1.append(df_Sample.columns.tolist())

        for index, row in df_Sample.iterrows():

            ws1.append([row.iloc[0], row.iloc[1], row.iloc[2], row.iloc[3], row.iloc[4], row.iloc[5], row.iloc[6], row.iloc[7], row.iloc[8], row.iloc[9], row.iloc[10], row.iloc[11]])
        ws1.merge_cells("A1:L1")
        # print(Lot,Sample_Type)
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 5
    ws1.column_dimensions['D'].width = 13
    ws1.column_dimensions['E'].width = 23
    ws1.column_dimensions['F'].width = 9
    ws1.column_dimensions['G'].width = 16
    ws1.column_dimensions['H'].width = 23
    ws1.column_dimensions['I'].width = 26
    ws1.column_dimensions['J'].width = 30
    ws1.column_dimensions['K'].width = 30
    ws1.column_dimensions['L'].width = 30

    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell_a in row:
            cell_a.border = border
            cell_a.alignment = alignment

def check_itemcode(ws5, HK_job,pbi_data,Insert_sheet):
    InsertionOrder = pd.read_excel(HK_job, sheet_name=Insert_sheet)
    hk1=InsertionOrder[~InsertionOrder["Unnamed: 1"].isna()]
    hk2=hk1[hk1["Lot"] != 'Ord']
    hk2 = hk2.assign(Ord=0.0)
    # hk2["Ord"]=hk2["Lot"].astype(float)
    # hk2.loc[:,"Ord"] = 0
    hk2.loc[:,"Ord"]=hk2["Lot"].astype(float)
    hk2.rename(columns={"Unnamed: 1":"Component"},inplace=True)


    pbi1_group=pbi_data.groupby(by=["Lot"])
    #Component,Lot
    # for lot,data in pbi1_group:
    ws5.append(["入封次序", "貨號", "異常情況"])
    for lot in pbi1_group.groups:
        data = pbi1_group.get_group((lot,))

        row_num=0
        data1 = data[["Ord", "Component", "Item Code"]]
        if lot in hk2.columns.tolist():
            hk3 = hk2[["Ord","Component", lot]]
            hk3=hk3.rename(columns={lot: "Item Code"})
            hk3=hk3.dropna(axis=0)
            check_mg1 = pd.merge(data1, hk3, on=["Ord", "Item Code"], how="outer")
            check_mg1.loc[check_mg1["Component_x"].isna(), "err"] = "PBI 輸出數據缺少資料"
            check_mg1.loc[check_mg1["Component_y"].isna(), "err"] = "HK 輸出數據缺少資料"
            check_mg2 =check_mg1.loc[~check_mg1["err"].isna(),["Ord","Item Code","err"]]
            # print(check_mg2)


            for index, row in check_mg2.iterrows():
                ws5.append([row["Ord"],row["Item Code"],row["err"]])
    check_count=ws5.max_row
    print("line:",check_count)
    if check_count==1:
        ws5.append(["", "", "全部貨號配對OK"])
    ws5.column_dimensions['B'].width = 15
    ws5.column_dimensions['C'].width = 30
        #
        # for index,row in data.iterrows():
        #     Ord=row["Ord"]
        #     ItemCode=row["Item Code"]
        #
        #     check=hk2.loc[(hk2["Ord"]==Ord) & (hk2[lot]==ItemCode)]
        #     if check.shape[0]==0:
        #         check_a = hk2.loc[(hk2["Ord"] == Ord)]
        #         check_b = hk2.loc[(hk2[lot] == ItemCode)]
        #         if check_a.shape[0] == 0:
        #             pass
        #         else:
        #
        #             print("Ord OK:",Ord,ItemCode,hk2.loc[hk2["Ord"]==Ord,lot].values[0])
        #         if check_b.shape[0] == 0:
        #             pass
        #         else:
        #
        #             print("ItemCode OK:", Ord, ItemCode, hk2.loc[hk2[lot] == ItemCode, "Ord"].values[0])
        #
        #     else:
        #         print("All OK: ",Ord,ItemCode)
        #     row_num=row_num+1



        # print(lot)

def Production_fanyi(Production_sheet:Sheet, sheet_src, master):

    def get_gsm(gsm):
        re4 = re.compile(r'\s?([0-9,\.]{1,8})\s{0,1}gsm', re.IGNORECASE)  # 取gsm
        s4 = re4.findall(gsm)
        if gsm == "House PSL Stock" or gsm == "House PSL Stock-MATTE":
            gsm_str = "190"
        elif s4:
            gsm_str = s4[0]
        else:
            gsm_str = constant.gsm_na

        return gsm_str

    def Laser_fanyi():
        """打印內容翻譯"""
        Special_Instructions = "" if row[32].data_type == 'n' else row[32].value
        Personalization_Details = "" if row[35].data_type == 'n' else row[35].value

        Personalization_Option = "" if row[33].data_type == 'n' else row[33].value
        Personalization_Option_cn = master.get_dict("personalizationOption", Personalization_Option)

        Imaging_Option = "" if row[49].data_type == 'n' else row[49].value

        Imaging_Option_cn = master.get_dict("imagingOption", Imaging_Option)

        # print(Personalization_Option_cn,Imaging_Option,Imaging_Option_cn)
        Laser_find = ""
        if (Special_Instructions + " " + Personalization_Details).lower().find("bule") > -1:
            Laser_find = "打印藍墨"
        Laser_cn = Personalization_Option_cn + " " + Imaging_Option_cn + ' ' + Laser_find
        return Laser_cn

    Production_sheet.range('A:A').insert(shift=None)
    dict_fix={
        "Perso":{True:"要打印",False:"","":""},
        "Personalization_Side": {"Front and Back": "雙面打印", "Front": "", "Back": "打印底面","":""},
        "Drives_Package": {True: "打印地址對窗口", False: "", "": ""},
        "Component_HandCode":{"Outer Box":"/左右各一個2\"透明圓形貼紙(有針綫)封口", "Sandwich Bags - Reusable":"/垂直中間需要壓壓線/每邊用1個1''圓形貼紙有針線封口","Address Label Sheet":"/啤半穿"},

              }
    row1=sheet_src[1]
    check_match = True
    for r in row1:
        mv=master.get_dict("ProductionReport_Head", r.column)
        if r.value != mv:
            check_match=False
            print(r"當前處理的Production Report 工作表字段與master主檔ProductionReport_Head工作表字段不對應, 請檢查.\nProduction Report 工作表字段:"+r.value+"\nmaster主檔ProductionReport_Head工作表字段:"+mv)
    max_row=sheet_src.max_row
    if check_match:
        for row in sheet_src.iter_rows(min_row=2):
            Component=row[3].value
            Item_Code_value = row[4].value
            Item_Code1="" if Item_Code_value is None else Item_Code_value[0]

            row_num=row[0].row
            #用紙
            gsm = "" if row[16].data_type == 'n' else row[16].value
            gsm_str = get_gsm(gsm)
            Stock_Type = "" if row[15].data_type == 'n' else row[15].value
            Stock_Type_cn = master.get_dict("stockType", Stock_Type)
            # Stock_Type_cn = Stock_Type_cn if Stock_Type_cn else constant.stockType_no_cn

            #折法
            Folding = "" if row[13].data_type == 'n' else row[13].value
            Folding_cn = master.get_dict("Folding", Folding)
            # Folding_cn = "折法沒找到!" if Folding_value == "" else Folding_value
            Folding_cn = "/" + Folding_cn if Folding_cn else Folding_cn

            #打印
            Perso = "" if row[9].data_type == 'n' else row[9].value
            Perso_cn = dict_fix.get("Perso").get(Perso)
            Perso_cn = "/" + Perso_cn if Perso_cn else ""

            #打印地址對窗口
            Drives_Package = "" if row[21].data_type == 'n' else row[21].value
            Drives_Package_cn = dict_fix.get("Drives_Package").get(Drives_Package)
            Drives_Package_cn = "/" + Drives_Package_cn if Drives_Package_cn else ""

            #打印面或底
            Personalization_Side = "" if row[34].data_type == 'n' else row[34].value
            Personalization_Side_cn = dict_fix.get("Personalization_Side").get(Personalization_Side)
            Personalization_Side_cn = "/" + Personalization_Side_cn if Personalization_Side_cn else ""
            #打印內容
            Laser_cn = Laser_fanyi()
            Laser_cn = "/" + Laser_cn if Laser_cn.strip() else ""

            #啤
            Perforation = "" if row[29].data_type == 'n' else row[29].value
            Perforation_cn = master.get_dict("perforation", Perforation)
            Perforation_cn = "/" + Perforation_cn if Perforation_cn else Perforation_cn

            if Component in ["OE - Paper","Reply Envelope"]:
                # 用紙|打印|打印內容
                Production_sheet.range(row_num,1).value=gsm_str+" "+ Stock_Type_cn +Perso_cn+Laser_cn

            elif Component in ["OE - Chipboard"]:
                # 用紙|
                Seal_Affix_Material = "" if row[56].data_type == 'n' else row[56].value
                Seal_Affix_Material_cn = master.get_dict("sealaffixedMaterial", Seal_Affix_Material)
                Seal_Affix_Material_cn = "/" + Seal_Affix_Material_cn if Seal_Affix_Material_cn else Seal_Affix_Material_cn
                Production_sheet.range(row_num,1).value=gsm_str+" "+ Stock_Type_cn +Seal_Affix_Material_cn
            elif Component in ["Outer Box","Sandwich Bags - Reusable"]:
                # 用紙|
                Production_sheet.range(row_num,1).value=gsm_str+" "+ Stock_Type_cn + dict_fix.get("Component_HandCode").get(Component)

            elif Component in ["Mailing Label"]:
                # 用紙|打印|  印色?
                Print_Option = "" if row[14].data_type == 'n' else row[14].value
                Print_Option_cn="/不用印刷" if Print_Option=="0/0" else "/印色:"+Print_Option
                Production_sheet.range(row_num, 1).value = gsm_str + " " + Stock_Type_cn + Perso_cn + Print_Option_cn

            elif Component in ["Letter","Reply Card","Letter/Reply Card","Letter/Reply Card/Check","Cover Letter","Cover Letter/Letter","Check","Check/Reply Card","Letter/Reply Card/Voucher","Reply Card-Snap Pack","Voucher/Reply Card","Petition/Reply Card","Survey/Reply Card"]:
                # 用紙|打印|打印內容|折法|啤
                Production_sheet.range(row_num, 1).value = gsm_str +" "+ Stock_Type_cn+Folding_cn + Perso_cn + Personalization_Side_cn +Drives_Package_cn+ Laser_cn+Perforation_cn
            elif Component in ["Card"]:
                # 用紙|打印|打印內容|折法 , 固定:單粉+ 貨號印粉面
                CD_cn = "/貨號印粉面" if Stock_Type_cn == '單粉' else ""
                Production_sheet.range(row_num, 1).value = gsm_str +" "+Stock_Type_cn + CD_cn + Folding_cn + Perso_cn + Laser_cn
            elif Component in ["Matching Envelope"]:
                # 用紙|打印
                Production_sheet.range(row_num, 1).value = gsm_str + " " + Stock_Type_cn + Perso_cn
            elif Component in ["Address Label Sheet"]:
                # 用紙|打印|折法|啤, 固定:啤半穿
                Production_sheet.range(row_num, 1).value = gsm_str +" "+ Stock_Type_cn+Folding_cn + Perso_cn +dict_fix.get("Component_HandCode").get(Component)+ Perforation_cn

            elif str(Component).startswith('NP') and Component!="NP-BK":
                com_sp = str(Component).index('-') + 1
                if str(Component[com_sp:]).strip().isnumeric():
                    NP_times = int(Component[com_sp:])
                    HandCode_NP="/1疊"+NP_times.__str__()+"張/NP+BK短頂邊pad頭塗膠水"
                # 用紙|打印|折法|啤, 固定:啤半穿
                Production_sheet.range(row_num, 1).value = gsm_str +" "+ Stock_Type_cn+ Perso_cn+HandCode_NP


            elif Item_Code1=='Q':
                Production_sheet.range(row_num, 1).value = Perso_cn + Laser_cn
            elif Component == None:
                pass
            else:
                Production_sheet.range(row_num, 1).value = gsm_str + " " + Stock_Type_cn + Folding_cn + Perso_cn
                # print("test",Component, gsm_str + " " + Stock_Type_cn + Perso_cn + Laser_cn)
                Production_sheet.range(row_num, 1).color = (255, 0, 255)

            #加顏色
            if str(Production_sheet.range(row_num, 1).value).find('主檔無中文')>-1 or str(Production_sheet.range(row_num, 1).value).find('找不到')>-1:
                Production_sheet.range(row_num, 1).color = (255, 255, 0)
        # Production_sheet.range((1,1),(max_row,1)).autofit()
        print("t")






wb = Workbook()
ws6 = wb.active
ws6.title = "Sample"
ws5 = wb.create_sheet("PBIData Match HKData")
ws4 = wb.create_sheet("Production_Report_ForCheck")
HK_filename,filename =read_lotus_Instruction("4096888-2")  #4096888-2, 4096786-2, 4097729-2, (4098746-2 這個沒A欄)    4097824-2,   OB test: 4097080-2
# master_Sample = pd.read_excel("./master/master.xlsx", sheet_name="Sample")
# master_stockType = pd.read_excel("./master/master.xlsx", sheet_name="stockType")
# master_Folding = pd.read_excel("./master/master.xlsx", sheet_name="Folding")
pbi_data = pd.read_excel(r"C:\Users\ITProg02\Downloads\data.xlsx", skiprows=2)

##
master_file="./master/master.xlsx"
master_sheet=["stockType","Folding","imagingOption","personalizationOption","numberOfPages","numSheet","perforation","sealaffixedMaterial","Sample","ProductionReport_Head"]
master1=master(master_file,master_sheet)
# master1.get_dict("Sample","Container")

##
from xlwings import Book,Sheet
# wb=xlwings.Book(filename)
# sheet0 : Sheet =wb.sheets[0]
if os.path.exists(HK_filename):
    wb_src = load_workbook(HK_filename, rich_text=False)
    Sample_check = False
    Insert_check = False
    for sn in wb_src.sheetnames:
        if str(sn).lower().startswith("sample"):
            Sample_check=True
            sample_sheet=sn
            sample(ws6, HK_filename, sample_sheet,master1)

        if str(sn).lower().startswith("insert"):
            Insert_check=True
            Insert_sheet=sn
            check_itemcode(ws5, HK_filename,pbi_data,Insert_sheet)

    if not Sample_check:
        ws6.append(["HK data 里沒有<Sample Details>表, 請檢查."])
        ws6.merge_cells("A1:L1")
    if not Insert_check:
        ws5.append(["HK data 里沒有<Insertion Order>表, 請檢查."])
        ws5.merge_cells("A1:L1")

    wb.save("temp.xlsx")
    print("file: temp.xlsx")

    shutil.copyfile(HK_filename, filename)
    wb_temp:Book = xlwings.Book("temp.xlsx")
    wb_P = xlwings.Book(filename)
    for sn1 in wb_P.sheet_names:
        if not str(sn1).lower().startswith("production"):
            wb_P.sheets[sn1].delete()
        else:
            wb_P.sheets[sn1].name = "Production_Report_ForCheck"
            Production_sheet = wb_P.sheets["Production_Report_ForCheck"]
            sheet_src=wb_src[sn1]
            if Sample_check:
                sample_sheet_copy:Sheet=wb_temp.sheets["Sample"]
                sample_sheet_copy.copy(after=wb_P.sheets[0])
            if Insert_check:
                Insert_sheet_copy:Sheet=wb_temp.sheets["PBIData Match HKData"]
                Insert_sheet_copy.copy(after=wb_P.sheets[0])
            wb_temp.close()
            Production_fanyi(Production_sheet, sheet_src, master1)

    # ws4=wb["Production_Report_ForCheck"]
    # wb.sheets.add("PBIData Match HKData", after=True)
    # wb.sheets.add("Sample", after=True)

    # wb.save(filename)
    wb_P.save(filename)

